import os
import oracledb
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import io
import base64
from dotenv import load_dotenv, find_dotenv

from utils.utils import get_embedding, summarize_image_to_text, summarize_text, get_image_embedding

_ = load_dotenv(find_dotenv())
oracledb.init_oracle_client()

UN = os.environ.get("UN")
PW = os.environ.get("PW")
DSN = os.environ.get("DSN")

def extract_tables_and_images(file_path, output_dir="images"):
  os.makedirs(output_dir, exist_ok=True)
  wb = load_workbook(file_path)
  e_wb = pd.ExcelFile(file_path)
  result = []

  for sheet_name in wb.sheetnames:
    # if not sheet_name in check_sheets:
    #   continue
    ws = wb[sheet_name]
    print(f"worksheet: {ws}")
    # --- 表の抽出 ---
    df = pd.read_excel(e_wb, sheet_name=sheet_name)
    markdown = df.to_markdown(index=False)
    print(f"Markdown Table:\n{markdown}")
    result.append({
      "type": "table",
      "sheet": sheet_name,
      "markdown": markdown,
      "raw": df
    })

    # --- 画像の抽出 ---
    for image in ws._images:
      if isinstance(image, XLImage):
        img_bytes = image._data()
        img = Image.open(io.BytesIO(img_bytes))

        # ファイル名決定
        img_name = f"{sheet_name}_{image.anchor._from.row}_{image.anchor._from.col}.png"
        img_path = os.path.join(output_dir, img_name)
        img.save(img_path)
        print(f"Image Path: {img_path}")

        result.append({
          "type": "image",
          "sheet": sheet_name,
          "image_path": img_path,
        })
  return result

def save_file_info(file_path: str):
  try:
    with oracledb.connect(user=UN, password=PW, dsn=DSN) as conn:
      cur = conn.cursor()
      file_name = file_path.split("/")[-1]
      file_id_var = cur.var(oracledb.NUMBER)
      sql = """INSERT INTO uploaded_files (file_name, file_path) VALUES (:1, :2) RETURNING id INTO :3"""
      data = [file_name, file_path, file_id_var]
      cur.execute(sql, data)
      file_id = file_id_var.getvalue()[0]
      conn.commit()
      return file_id
  except oracledb.DatabaseError as e:
    error, = e.args
    print(f"Error at save_file_info")
    print(f"Oracle error code: {error.code}")
    print(f"Oracle error message: {error.message}")
    return None

def save_docs_content(file_id, markdown: str, summary: str, embedding: list):
  sql = """
          INSERT INTO docs_contents (file_id, summary, embedding, markdown)
          VALUES (:file_id, :summary, :embedding, :markdown)
          """
  params = {
    "file_id": file_id,
    "summary": summary,
    "embedding": str(embedding),
    "markdown": markdown,
  }
  try:
    with oracledb.connect(user=UN, password=PW, dsn=DSN) as conn:
      with conn.cursor() as cursor:
        cursor.execute(sql, params)
        print(f"Success insert {file_id} into docs_contents")
        conn.commit()
  except oracledb.DatabaseError as e:
    error, = e.args
    print(f"Error at save_content")
    print(f"Oracle error code: {error.code}")
    print(f"Oracle error message: {error.message}")
  except Exception as e:
    print(f"Error:save_docs_content: {e}")
    return None

def save_image_content(file_id: str, image_path: str, summary: str, embedding: list):
  sql = """
        INSERT INTO image_contents (file_id, image_path, summary, embedding, image_blob)
        VALUES (:file_id, :image_path, :summary, :embedding, empty_blob())
        returning image_blob into :blobdata
      """
  
  try:
    with oracledb.connect(user=UN, password=PW, dsn=DSN) as conn:
      with conn.cursor() as cursor:
        blobdata = cursor.var(oracledb.DB_TYPE_BLOB)
        params = {
          'file_id': file_id,
          'image_path': image_path,
          'summary': summary,
          'embedding': str(embedding),
          'blobdata': blobdata
        }
        cursor.execute(sql, params)
        blob, = blobdata.getvalue()
        offset = 1
        bufsize = 65536
        with open(image_path, 'rb') as f:
            while True:
                data = f.read(bufsize)
                if data:
                    blob.write(data, offset)
                if len(data) < bufsize:
                    break
                offset += bufsize
        
        print(f"Success insert {file_id} into image_contents")
      conn.commit()
  except oracledb.DatabaseError as e:
    error, = e.args
    print(f"Error at save_content")
    print(f"Oracle error code: {error.code}")
    print(f"Oracle error message: {error.message}")
  except Exception as e:
    print(f"Error:save_image_content: {e}")
    return None


def summarize_to_db_and_upload_image(image_path):
  with open(image_path, "rb") as img_file:
    image_lob = img_file.read()
    print(f"type: {type(image_lob)}")
  
  print(f"image_path: {image_path}")
  image_summary = summarize_image_to_text(image_path)
  print(f"Image Summary: {image_summary}")
  # image_embedding = get_embedding(image_summary)
  image_embedding = get_image_embedding(image_path)
  
  image_id = save_file_info(image_path)
  save_image_content(image_id, image_path, image_summary, image_embedding)
  return {"summary": image_summary}


def process_excel_with_images(file_path):
  file_id = save_file_info(file_path)
  contents = extract_tables_and_images(file_path)

  for content in contents:
    if content["type"] == "table":
      print(f"content_type:table: {content["type"]}")
      summary = summarize_text(content["markdown"])
      embedding = get_embedding(summary)
      save_docs_content(file_id, content["markdown"], summary, embedding)

    elif content["type"] == "image":
      print(f"content_type:image: {content["type"]}")
      res = summarize_to_db_and_upload_image(
        image_path=content["image_path"]
        )


if __name__ == "__main__":
  process_excel_with_images("./data/fy25q3-supplemental.xlsx")
  summarize_to_db_and_upload_image(image_path="./images/net_income.png")
  summarize_to_db_and_upload_image(image_path="./images/revenue.png")
