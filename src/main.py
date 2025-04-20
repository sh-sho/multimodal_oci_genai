import os
import oci
from oci.generative_ai_inference.models import (
    ChatDetails, 
    CohereChatRequest, 
    OnDemandServingMode, 
    CohereTool, 
    CohereParameterDefinition,
    LlamaLlmInferenceRequest,
    GenericChatRequest,
    TextContent,
    Message,
    BaseChatRequest
)
from oci.generative_ai_inference import GenerativeAiInferenceClient
import oracledb
import requests
from langchain_community.chat_models import ChatOCIGenAI
from langchain_core.messages import HumanMessage
from langchain_core.prompts import PromptTemplate
from langchain.schema.output_parser import StrOutputParser
from langchain.schema.runnable import RunnablePassthrough
from langchain_community.llms.oci_generative_ai import OCIGenAI
from langchain.chains.llm import LLMChain


from langchain_core.documents import Document
from langchain_community.embeddings import OCIGenAIEmbeddings
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import io
import os
import base64
from tabulate import tabulate
from dotenv import load_dotenv, find_dotenv
_ = load_dotenv(find_dotenv())
oracledb.init_oracle_client()

UN = os.environ.get("UN")
PW = os.environ.get("PW")
DSN = os.environ.get("DSN")

OCI_CONFIG_FILE = "~/.oci/config"
OCI_COMPARTMENT_ID = os.getenv("OCI_COMPARTMENT_ID")
OCI_GENAI_ENDPOINT = os.getenv("OCI_GENAI_ENDPOINT")
config = oci.config.from_file(file_location=OCI_CONFIG_FILE, profile_name="OSAKA")
generative_ai_inference_client = GenerativeAiInferenceClient(config, sevice_endpoint=OCI_GENAI_ENDPOINT)
model_id="ocid1.generativeaimodel.oc1.ap-osaka-1.amaaaaaask7dceyayobenuynf4om42tyrr7scxwijpzwfajc6i6w5wjcmjbq"

preamble = """
## 回答のスタイル
日本語で回答してください。
質問に対してできる限り詳細な回答をしてください。
"""
input_text = "Oracleのクラウドについて教えてください"

# chat_detail = ChatDetails(
#     chat_request=CohereChatRequest(
#         preamble_override=preamble,
#         message=input_text,
#         max_tokens=500,
#         is_force_single_step=False,
#         ),
#     compartment_id=OCI_COMPARTMENT_ID,
#     serving_mode=OnDemandServingMode(
#         model_id=model_id
#     ))
# chat_response = generative_ai_inference_client.chat(chat_detail)

# res_chat = chat_response.data.chat_response
# print(f"Cohere Response: {res_chat}")



# generative_ai_inference_client = GenerativeAiInferenceClient(config=config, service_endpoint=OCI_GENAI_ENDPOINT, retry_strategy=oci.retry.NoneRetryStrategy(), timeout=(10,240))
# chat_detail = ChatDetails()

# content = TextContent()
# content.text = input_text
# message = Message()
# message.role = "USER"
# message.content = [content]

# chat_request = GenericChatRequest()
# chat_request.api_format = BaseChatRequest.API_FORMAT_GENERIC
# chat_request.messages = [message]
# chat_request.max_tokens = 600
# chat_request.temperature = 1
# chat_request.frequency_penalty = 0
# chat_request.presence_penalty = 0
# chat_request.top_p = 0.75

# chat_detail.serving_mode = OnDemandServingMode(model_id="ocid1.generativeaimodel.oc1.ap-osaka-1.amaaaaaask7dceyac2pavq6pya22whj4gvy5l7mpdyrlm646dt7n3cppfxcq")
# chat_detail.chat_request = chat_request
# chat_detail.compartment_id = OCI_COMPARTMENT_ID

# chat_response = generative_ai_inference_client.chat(chat_detail)
# res_chat = chat_response.data.chat_response
# print(f"Llama Response: {res_chat}")




# AUTH_TYPE = "API_KEY"
# CONFIG_PROFILE = "OSAKA"

# prompt="Oracleのクラウドについて教えてください"

# # Service endpoint
# endpoint = "https://inference.generativeai.ap-osaka-1.oci.oraclecloud.com"

# # initialize interface
# chat = ChatOCIGenAI(
#   model_id="ocid1.generativeaimodel.oc1.ap-osaka-1.amaaaaaask7dceyac2pavq6pya22whj4gvy5l7mpdyrlm646dt7n3cppfxcq",
#   service_endpoint=endpoint,
#   compartment_id=OCI_COMPARTMENT_ID,
#   provider="meta",
#   model_kwargs={
#     "temperature": 1,
#     "max_tokens": 600,
#     "frequency_penalty": 0,
#     "presence_penalty": 0,
#     "top_p": 0.75
#   },
#   auth_type=AUTH_TYPE,
#   auth_profile=CONFIG_PROFILE
# )

# messages = [
#   HumanMessage(content=prompt),
# ]

# response = chat.invoke(messages)
# res_chat = response.content
# print(f"ChatOCIGenAI Response: {res_chat}")




def extract_tables_and_images(file_path, output_dir="images"):
  os.makedirs(output_dir, exist_ok=True)
  wb = load_workbook(file_path)
  result = []

  for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # --- 表の抽出 ---
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    if not df.empty:
      markdown = tabulate(df.head(10), headers='keys', tablefmt='github')
      result.append({
        "type": "table",
        "sheet": sheet_name,
        "markdown": markdown,
        "raw": df
      })

    # --- 画像の抽出 ---
    for image in ws._images:
      if isinstance(image, XLImage):
        img_bytes = image._data()
        img = Image.open(io.BytesIO(img_bytes))

        # ファイル名決定
        img_name = f"{sheet_name}_{image.anchor._from.row}_{image.anchor._from.col}.png"
        img_path = os.path.join(output_dir, img_name)
        img.save(img_path)

        # Base64に変換（任意）
        with open(img_path, "rb") as f:
          base64_str = base64.b64encode(f.read()).decode("utf-8")
          data_url = f"data:image/png;base64,{base64_str}"

        result.append({
          "type": "image",
          "sheet": sheet_name,
          "image_path": img_path,
          "image_url": data_url
        })
  return result

def save_file_info(file_path):
  try:
    with oracledb.connect(user=UN, password=PW, dsn=DSN) as conn:
      cur = conn.cursor()
      file_name = file_path.split("/")[-1]
      file_id_var = cur.var(oracledb.NUMBER)
      sql = """INSERT INTO uploaded_files (file_name, file_path) VALUES (:1, :2) RETURNING id INTO :3"""
      data = [file_name, file_path, file_id_var]
      cur.execute(sql, data)
      file_id = file_id_var.getvalue()[0]
      conn.commit()
      return file_id
  except oracledb.DatabaseError as e:
    error, = e.args
    print(f"Error at save_file_info")
    print(f"Oracle error code: {error.code}")
    print(f"Oracle error message: {error.message}")
    return None

def save_content(file_id, markdown, summary, embedding, content_type="table"):
  try:
    with oracledb.connect(user=UN, password=PW, dsn=DSN) as conn:
      cur = conn.cursor()
      embedding_str = ",".join(map(str, embedding))
      cur.execute("""
        INSERT INTO file_contents (file_id, content_type, markdown, summary, embedding)
        VALUES (:1, :2, :3, :4, :5)
      """, [file_id, content_type, markdown, summary, embedding_str])
      conn.commit()
  except oracledb.DatabaseError as e:
    error, = e.args
    print(f"Error at save_content")
    print(f"Oracle error code: {error.code}")
    print(f"Oracle error message: {error.message}")

def summarize_text(text):
  prompt = PromptTemplate(
      input_variables=["text"],
      template="以下の内容を要約してください:\n{text}"
  )
  llm = ChatOCIGenAI(
    model_id="cohere.command-r-08-2024",
    service_endpoint="https://inference.generativeai.ap-osaka-1.oci.oraclecloud.com",
    compartment_id=OCI_COMPARTMENT_ID,
    model_kwargs={"temperature": 0.7, "max_tokens": 500},
    )
  # chain = LLMChain(llm=llm, prompt=prompt)
  chain = ({"text": RunnablePassthrough() } 
           | prompt 
           | llm 
           | StrOutputParser()
  )
  result = chain.invoke(text)
  return result

def summarize_image_to_text(text):
  prompt = PromptTemplate(
      input_variables=["text"],
      template="以下の内容を要約してください:\n{text}"
  )
  llm = ChatOCIGenAI(
    model_id="cohere.command-r-08-2024",
    service_endpoint="https://inference.generativeai.ap-osaka-1.oci.oraclecloud.com",
    compartment_id=OCI_COMPARTMENT_ID,
    model_kwargs={"temperature": 0.7, "max_tokens": 500},
    )
  chain = ({"text": RunnablePassthrough() } 
           | prompt 
           | llm 
           | StrOutputParser()
  )
  result = chain.invoke(text)
  return result

def get_embedding(text):
  embeddings = OCIGenAIEmbeddings(
    model_id="cohere.embed-multilingual-v3.0",
    service_endpoint="https://inference.generativeai.ap-osaka-1.oci.oraclecloud.com",
    compartment_id=OCI_COMPARTMENT_ID,
  )
  return embeddings.embed_query(text)

def process_excel_with_images(file_path):
  file_id = save_file_info(file_path)
  contents = extract_tables_and_images(file_path)

  for content in contents:
    if content["type"] == "table":
      summary = summarize_text(content["markdown"])
      embedding = get_embedding(summary)
      save_content(file_id, content["markdown"], summary, embedding, content_type="table")

    elif content["type"] == "image":
      summary = summarize_image_to_text("この画像の内容を要約してください。")
      embedding = get_embedding(summary)
      save_content(file_id, content["image_url"], summary, embedding, content_type="image")


if __name__ == "__main__":
  process_excel_with_images("./data/sample_sales_infos.xlsx")